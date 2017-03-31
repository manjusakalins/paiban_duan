#coding=utf-8
import os
import pickle
import codecs
import sys
import xlrd
from openpyxl import load_workbook
import xlsxwriter
import re
import operator
import random
import cairo
import calendar
import time
from datetime import date

## data struct ###
## everyone have {"yb":time, "jqb":time, "name":"rname", yb_timelist[], jqb_timelist[]}
# glist ={"name":struct, "name":struct}

g_list={};
g_st_list=[];
g_name_list=[];
g_num_mon=3;
g_start_mon=4;
g_cur_date_int=0;
g_cur_flag="jqb";

#def is_china_holiday(year, mon, day):
#if rs.cell("H8").fill == rs.cell("H4").fill :
#                    print "nn"
#                else:
#                    print "not none"

def lhs_get_another_flag(flag):
    if flag == "yb":
        return "jqb"
    else:
        return "yb"

def lhs_get_remain_time(cur_one, flag):
    if not cur_one.has_key(flag):
        return 0;
    did = len(cur_one[flag+"l"]);
    return cur_one[flag] - did;

def lhs_get_cur_one_weight(one):
    
    global g_cur_date_int;
    remain_mon = g_num_mon - (date.fromordinal(g_cur_date_int).month - g_start_mon);#3 2 1
    remain_work = lhs_get_remain_time(one, g_cur_flag);
    #print remain_mon
    #print remain_work;
    tmps=float(float(remain_work)/float(remain_mon));
    #print round(tmps)
    return round(tmps);

def lhs_load_xlsx_jiaqiangban(fname, sheet_name, btype, start_col, end_col, time_row, start_row, end_row):
    end_col = end_col+1;
    end_row = end_row+1;
    print fname
    rb = load_workbook(fname);
    print rb.get_sheet_names();

    input_f=fname;
    input_dir=input_f[:input_f.rfind('/')]
    print input_dir

    for cur_sheet in rb.get_sheet_names():
        if cur_sheet == sheet_name:
            print rb[cur_sheet]
            rs = rb[cur_sheet]

            current_time = 0;
            for idx in range(start_col, end_col):
                current_time = int(rs.cell(row=time_row, column=idx).value);
                print current_time;
                for jdx in range(start_row, end_row):
                    cur_name = rs.cell(row=jdx, column=idx).value;
                    if cur_name is None:
                        continue;
                    print cur_name
                    if g_list.has_key(cur_name):
                        cur_st=g_list[cur_name]
                    else:
                        cur_st={"ybl":[], "jqbl":[]};
                    cur_st["name"]=cur_name
                    cur_st[btype]=current_time;
                    g_list[cur_name]=cur_st;

                #print rs.cell("H4").fill;
                #print "aa"
                #print rs.cell("H8").fill;
                #if rs.cell("H32").fill == rs.cell("H4").fill :
                #    print "nn"
                #else:
                #    print "not none"



def lhs_check_one_can_work(idx, b_flag, date_int):
    #check it.
    #print date.fromordinal(date_int)
    cur_one = g_list[g_name_list[idx]]

    
    #first check whether has b_flag
    if not cur_one.has_key(b_flag):
        return 0;
    # ban all done
    if lhs_get_remain_time(cur_one, b_flag) <= 0:
        return 0;


    # a totally new one, just return 1
    if (len(cur_one["ybl"]) == 0) and (len(cur_one["jqbl"]) == 0) :
        return 1;


    if b_flag == "jqb":
        #we make it mean
        if len(cur_one["jqbl"]) > 0:
                last_jqb = cur_one["jqbl"][len(cur_one["jqbl"])-1];
                if date_int - last_jqb < 10: #6time/3mon, so min is 10 day
                    return 0;

        if len(cur_one["ybl"]) > 0:
            last_yb = cur_one["ybl"][len(cur_one["ybl"])-1];
            if date_int == last_yb: #not the same day:
                return 0;
            if (date_int - last_yb) > 4:#at least 4 day:
                return 1;

        else:
            #no yb did:
            return 1;

    if b_flag == "yb":
        #not the same day:
        if len(cur_one["ybl"]) > 0:
            last_yb = cur_one["ybl"][len(cur_one["ybl"])-1];
            if date_int == last_yb:
                return 0;
            #at least 4 day:
            if (date_int - last_yb) > 10:
                return 1;
        else:
            #no yb did:
            if len(cur_one["jqbl"]) > 0:
                last_jqb = cur_one["jqbl"][len(cur_one["jqbl"])-1];
                if date_int - last_jqb > 4:
                    return 1;
            return 0;


def lhs_check_work_and_set(idx, b_flag, date_int):
    if lhs_check_one_can_work(idx, b_flag, date_int) == 1:
        #g_list[g_name_list[idx]][b_flag] = g_list[g_name_list[idx]][b_flag] - 1
        g_list[g_name_list[idx]][b_flag+"l"].append(date_int)
        return 1;
    else:
        return 0;


def lhs_arange_one_day(date_int, is_hol):
    global g_cur_date_int;
    global g_cur_flag;
    t_num = len(g_name_list);
    #print t_num;
    jqb = 4;
    yb = 2;
    if is_hol == 1:
        jqb = 3;
    g_cur_date_int = date_int;


    print "arrange day work"
    tmp_st_list=[];
    g_cur_flag = "jqb";
    for cur_st in g_list:
        tmp_st_list.append(g_list[cur_st]);
    sorted_list = sorted(tmp_st_list, key=lhs_get_cur_one_weight, reverse = True);
    for idx in sorted_list:
        print lhs_get_remain_time(idx,g_cur_flag);

    ##jqb
    for sidx in sorted_list:
        curi = sidx["idx"];
        if jqb <= 0:
            break;
        can_work = lhs_check_work_and_set(curi, g_cur_flag, date_int);
        if (can_work == 1):
            print g_name_list[curi]
            print("do in jqb: ", date.fromordinal(date_int),  g_list[g_name_list[curi]][g_cur_flag])
            jqb = jqb - 1;
    print ("remain ", g_cur_flag, jqb)

    while jqb > 0:
        jqb = 1;

    print "arrange night work"
    tmp_st_list=[];
    g_cur_flag = "yb";
    for cur_st in g_list:
        tmp_st_list.append(g_list[cur_st]);
    sorted_list = sorted(tmp_st_list, key=lhs_get_cur_one_weight, reverse = True);
    for idx in sorted_list:
        print lhs_get_remain_time(idx,g_cur_flag);

    #yb
    for sidx in sorted_list:
        curi = sidx["idx"];
        if yb <= 0:
            break;
        can_work = lhs_check_work_and_set(curi, g_cur_flag, date_int);
        if (can_work == 1):
            print g_name_list[curi]
            print("do in ", g_cur_flag, date.fromordinal(date_int),  g_list[g_name_list[curi]][g_cur_flag])
            yb = yb - 1;

    print ("remain ", g_cur_flag, yb)






def lhs_start_fill_blank(fname, sheet_name, holiday_pos, day_col, start_col, end_col, start_row, end_row, st_int):
    end_col = end_col+1;
    end_row = end_row+1;
    print fname
    rb = load_workbook(fname);
    print rb.get_sheet_names();

    input_f=fname;
    input_dir=input_f[:input_f.rfind('/')]
    print input_dir

    for cur_sheet in rb.get_sheet_names():
        if cur_sheet == sheet_name:
            print rb[cur_sheet]
            rs = rb[cur_sheet]
            holiday_fill = rs.cell(holiday_pos).fill;
            hol_flag=0;
            for idx in range(start_row, end_row):
                cur_fill = rs.cell(row=idx, column=day_col).fill;
                #print cur_fill;
                if cur_fill == holiday_fill:
                    print "holiday"
                    hol_flag = 1;
                ############ arange one day ############
                lhs_arange_one_day(st_int+idx-start_row, hol_flag)







lhs_load_xlsx_jiaqiangban("/home/manjusaka/all_codes/new_my_code/in_github/duan_paiban/input.xlsx", "Sheet1", "jqb", 2,7, 4,5,27);
lhs_load_xlsx_jiaqiangban("/home/manjusaka/all_codes/new_my_code/in_github/duan_paiban/input.xlsx", "Sheet1", "yb", 2,6, 30,31,53);
print g_list
name_idx=0;
for cur_st in g_list:
    print cur_st
    g_list[cur_st]["idx"]=name_idx;
    g_name_list.append(cur_st);
    g_st_list.append(g_list[cur_st]);
    name_idx = name_idx + 1;

    
print calendar.weekday(2017, 4, 15)
print calendar.weekday(2017, 4, 16)
print date(date.today().year, 4, 2)
st_date = date(date.today().year, 4, 1)
print calendar.weekday(st_date.year, st_date.month, st_date.day)

st_int=st_date.toordinal();
print date.fromordinal(st_int+1)
lhs_start_fill_blank("/home/manjusaka/all_codes/new_my_code/in_github/duan_paiban/input.xlsx", "Sheet1", "H4", 9, 10, 15, 3, 93, st_int);

